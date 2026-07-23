from __future__ import annotations

import html
from io import BytesIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


NAVY = "17324D"
TEAL = "187B7A"
LIGHT = "EAF1F7"
GREEN = "D9EAD3"
RED = "F4CCCC"
WHITE = "FFFFFF"


def _html_list(items) -> str:
    values = items if isinstance(items, list) else ([items] if items else [])
    return "<ul>" + "".join(f"<li>{html.escape(str(item))}</li>" for item in values) + "</ul>"


def _report_html(title: str, body: str) -> bytes:
    document = f"""<!doctype html>
<html lang="zh-CN"><head><meta charset="utf-8"><title>{html.escape(title)}</title>
<style>
body{{font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif;margin:36px;color:#17324d}}
h1{{border-bottom:4px solid #187b7a;padding-bottom:12px}} h2{{margin-top:30px;color:#155e63}}
.note{{background:#eef6f5;padding:14px;border-left:5px solid #187b7a}}
table{{border-collapse:collapse;width:100%;font-size:13px;margin:12px 0 24px}}
th{{background:#17324d;color:white}} th,td{{border:1px solid #d6dee6;padding:8px;text-align:left;vertical-align:top}}
tr:nth-child(even){{background:#f6f8fb}} li{{margin:6px 0}}
</style></head><body><h1>{html.escape(title)}</h1>{body}</body></html>"""
    return document.encode("utf-8")


def _style_sheet(ws, freeze: str = "A2") -> None:
    ws.freeze_panes = freeze
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26
    for column_cells in ws.columns:
        values = [str(cell.value or "") for cell in list(column_cells)[:200]]
        width = min(max(max(map(len, values), default=8) + 2, 10), 44)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _write_frame(wb: Workbook, title: str, df: pd.DataFrame):
    ws = wb.create_sheet(title[:31])
    ws.append(list(df.columns))
    safe_df = df.astype(object).where(pd.notna(df), None)
    for row in safe_df.itertuples(index=False, name=None):
        ws.append(list(row))
    _style_sheet(ws)
    for idx, col in enumerate(df.columns, start=1):
        if "url" in str(col).lower() or "link" in str(col).lower():
            for row in range(2, ws.max_row + 1):
                value = ws.cell(row, idx).value
                if isinstance(value, str) and value.startswith("http"):
                    ws.cell(row, idx).hyperlink = value
                    ws.cell(row, idx).style = "Hyperlink"
    return ws


def _bytes(wb: Workbook) -> bytes:
    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def template_workbook(columns: list[str], sheet_name: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    ws.append(columns)
    _style_sheet(ws)
    return _bytes(wb)


def dataframe_to_csv_bytes(df: pd.DataFrame) -> bytes:
    return df.to_csv(index=False).encode("utf-8-sig")


def build_bsr_workbook(
    products: pd.DataFrame,
    reviews: pd.DataFrame,
    analysis: dict,
    previous_products: pd.DataFrame,
    report_title: str,
) -> bytes:
    data = products.copy()
    if not previous_products.empty and "asin" in previous_products and "rank" in previous_products:
        previous = previous_products[["asin", "rank"]].rename(columns={"rank": "previous_rank"})
        data = data.merge(previous, on="asin", how="left")
        data["rank_shift"] = data["previous_rank"] - data["rank"]
    else:
        data["previous_rank"] = pd.NA
        data["rank_shift"] = pd.NA
    data["estimated_revenue"] = data["price"] * data["bought_past_month"]
    data["data_note"] = "销量/销售额为页面线索估算，非Seller Central真实数据"

    wb = Workbook()
    summary = wb.active
    summary.title = "Executive Summary"
    summary.append(["模块", "内容"])
    summary.append(["报告", report_title])
    summary.append(["市场格局", analysis.get("market_landscape", "尚未生成")])
    pain = analysis.get("pain_points", [])
    pain_text = "\n".join(
        (
            f"{item.get('pain_point')}: {item.get('mention_rate', 0)}%"
            if isinstance(item, dict)
            else str(item)
        )
        for item in pain
    )
    summary.append(["VOC痛点", pain_text or "暂无足够评论"])
    selling = analysis.get("selling_points", [])
    selling_text = "\n".join(
        (
            f"{item.get('selling_point')}: {item.get('mention_rate', 0)}%"
            if isinstance(item, dict)
            else str(item)
        )
        for item in selling
    )
    summary.append(["核心卖点", selling_text or "暂无足够好评"])
    summary.append(["品牌/技术对比", analysis.get("brand_comparison", "") or "暂无"])
    positioning = analysis.get("vacmaster_positioning", {})
    if isinstance(positioning, dict):
        positioning_text = "\n".join(f"{key}: {value}" for key, value in positioning.items())
    else:
        positioning_text = str(positioning or "")
    summary.append(["Vacmaster定位", positioning_text or "当前样本不足"])
    summary.append(
        ["产品/价格机会", "\n".join(map(str, analysis.get("opportunity_gaps", []))) or "暂无"]
    )
    summary.append(["对Vacmaster建议", "\n".join(analysis.get("recommendations", []))])
    summary.append(
        [
            "数据说明",
            "榜单、详情、价格、Coupon和评论可能受Amazon页面限制；"
            "评论正文需通过上传或合规数据源提供，估算值已明确标记。",
        ]
    )
    _style_sheet(summary)
    summary.column_dimensions["A"].width = 24
    summary.column_dimensions["B"].width = 100
    summary.row_dimensions[3].height = 70
    summary.row_dimensions[4].height = 80
    summary.row_dimensions[5].height = 80
    summary.row_dimensions[6].height = 90
    summary.row_dimensions[7].height = 100
    summary.row_dimensions[8].height = 100
    summary.row_dimensions[9].height = 100

    _write_frame(wb, "Top 50 SKUs Dataset", data)
    price_bands = analysis.get("price_band_analysis", [])
    if price_bands:
        _write_frame(wb, "Price Band Analysis", pd.DataFrame(price_bands))
    benchmark = analysis.get("brand_benchmark", [])
    if benchmark:
        _write_frame(wb, "Brand Benchmark", pd.DataFrame(benchmark))
    capability = analysis.get("capability_comparison", [])
    if capability:
        _write_frame(wb, "Capability Comparison", pd.DataFrame(capability))
    if not reviews.empty:
        _write_frame(wb, "Review Dataset", reviews)
    return _bytes(wb)


def build_bsr_html(products: pd.DataFrame, analysis: dict, report_title: str) -> bytes:
    positioning = analysis.get("vacmaster_positioning", {})
    if isinstance(positioning, dict):
        positioning_html = "<table><tr><th>分析项</th><th>结论</th></tr>" + "".join(
            f"<tr><td>{html.escape(str(key))}</td><td>{html.escape(str(value))}</td></tr>"
            for key, value in positioning.items()
        ) + "</table>"
    else:
        positioning_html = f"<p>{html.escape(str(positioning or '暂无'))}</p>"
    sections = [
        f"<h2>市场格局</h2><p>{html.escape(str(analysis.get('market_landscape', '暂无')))}</p>",
        "<h2>Vacmaster定价与能力定位</h2>" + positioning_html,
        "<h2>产品与价格机会</h2>" + _html_list(analysis.get("opportunity_gaps", [])),
        "<h2>应对建议</h2>" + _html_list(analysis.get("recommendations", [])),
    ]
    for heading, key in [
        ("价格带分析", "price_band_analysis"),
        ("品牌基准", "brand_benchmark"),
        ("能力差异", "capability_comparison"),
    ]:
        rows = analysis.get(key, [])
        if rows:
            sections.append(
                f"<h2>{heading}</h2>"
                + pd.DataFrame(rows).to_html(index=False, escape=True, border=0)
            )
    display_columns = [
        column
        for column in [
            "rank",
            "asin",
            "brand",
            "title",
            "price",
            "model",
            "capacity",
            "horsepower",
            "airflow",
            "suction",
            "clean_tank_capacity",
            "dirty_tank_capacity",
            "heating_or_steam",
            "accessories",
            "warranty",
            "detail_status",
        ]
        if column in products
    ]
    sections.append(
        "<h2>Top产品明细</h2>"
        + products[display_columns].to_html(index=False, escape=True, border=0)
    )
    sections.append(
        f"<p class='note'>分析方式：{html.escape(str(analysis.get('analysis_mode', '本地规则')))}。"
        "页面月购买量及推算销售额不等于Seller Central真实销量。</p>"
    )
    return _report_html(report_title, "".join(sections))


def build_competitor_workbook(intel: pd.DataFrame, summary: dict | None = None) -> bytes:
    wb = Workbook()
    if summary:
        ws = wb.active
        ws.title = "Executive Summary"
        ws.append(["模块", "内容"])
        mapping = [
            ("管理层摘要", "executive_summary"),
            ("核心技术规格与卖点", "core_technology_and_selling_points"),
            ("定价与市场策略", "pricing_and_market_strategy"),
            ("对CLEVA竞争影响", "competitive_impact_on_cleva"),
            ("应对方案", "response_plan"),
            ("重点关注", "priority_watchlist"),
            ("分析方式", "analysis_mode"),
        ]
        for label, key in mapping:
            value = summary.get(key, "")
            ws.append([label, "\n".join(map(str, value)) if isinstance(value, list) else str(value)])
        _style_sheet(ws)
        ws.column_dimensions["A"].width = 26
        ws.column_dimensions["B"].width = 110
    else:
        wb.remove(wb.active)
    _write_frame(wb, "Competitor Intelligence", intel)
    return _bytes(wb)


def build_competitor_html(intel: pd.DataFrame, summary: dict) -> bytes:
    body = (
        f"<h2>管理层摘要</h2><p>{html.escape(str(summary.get('executive_summary', '暂无')))}</p>"
        "<h2>核心技术规格与卖点</h2>"
        + _html_list(summary.get("core_technology_and_selling_points", []))
        + "<h2>定价与市场策略</h2>"
        + _html_list(summary.get("pricing_and_market_strategy", []))
        + "<h2>对CLEVA（Vacmaster/Lawnmaster）的竞争影响</h2>"
        + _html_list(summary.get("competitive_impact_on_cleva", []))
        + "<h2>应对方案</h2>"
        + _html_list(summary.get("response_plan", []))
        + "<h2>重点关注清单</h2>"
        + _html_list(summary.get("priority_watchlist", []))
        + "<h2>情报明细</h2>"
        + intel.to_html(index=False, escape=True, border=0)
        + f"<p class='note'>分析方式：{html.escape(str(summary.get('analysis_mode', '本地汇总')))}。"
        "候选情报及AI提取结果均需回到原始来源人工复核。</p>"
    )
    return _report_html("CLEVA全球竞品新品情报报告", body)


def build_sales_workbook(result: pd.DataFrame, as_of) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    ws = _write_frame(wb, "SKU Sales Analysis", result)
    percent_columns = [
        "wow_units", "wow_revenue", "mom_units", "mom_revenue",
        "qoq_units", "qoq_revenue", "hoh_units", "hoh_revenue",
        "yoy_units", "yoy_revenue", "ctr", "cvr", "acos",
    ]
    for name in percent_columns:
        if name not in result:
            continue
        idx = result.columns.get_loc(name) + 1
        letter = get_column_letter(idx)
        for row in range(2, ws.max_row + 1):
            ws.cell(row, idx).number_format = "0.0%"
        ws.conditional_formatting.add(
            f"{letter}2:{letter}{ws.max_row}",
            CellIsRule(operator="greaterThan", formula=["0"], fill=PatternFill("solid", fgColor=GREEN)),
        )
        ws.conditional_formatting.add(
            f"{letter}2:{letter}{ws.max_row}",
            CellIsRule(operator="lessThan", formula=["0"], fill=PatternFill("solid", fgColor=RED)),
        )
    for name in ["revenue_current", "price", "spend", "ad_sales"]:
        if name in result:
            idx = result.columns.get_loc(name) + 1
            for row in range(2, ws.max_row + 1):
                ws.cell(row, idx).number_format = '$#,##0.00'
    return _bytes(wb)
